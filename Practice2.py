#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import load_workbook


# In[2]:


read_workbook = load_workbook("./natepann.xlsx")


# In[3]:


read_cell = read_workbook.active


# In[4]:


result_list = []


# In[10]:


for i in range(1,101):
    result_list.append(read_cell.cell(i,1).value)


# In[11]:


print(result_list)


# In[ ]:




